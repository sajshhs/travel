"""Iteration 115 — IMPOR MASTER DATA (multi-sheet Excel migration).
Covers: GET /api/data/import/template.xlsx, /sheets, POST /preview, POST /commit (upsert/insert_only/sheets filter),
GET /api/data/import/history, RBAC, validation errors, idempotency, snapshot backup.
Run serially: pytest /app/backend/tests/test_master_import_i115.py -v -p no:randomly -n 0
"""
import io
import os
import re
from pathlib import Path

import pytest
import requests
from dotenv import dotenv_values
from openpyxl import load_workbook
from pymongo import MongoClient

fe = dotenv_values("/app/frontend/.env")
be = dotenv_values("/app/backend/.env")
BASE_URL = (os.environ.get("REACT_APP_BACKEND_URL") or fe.get("REACT_APP_BACKEND_URL") or "").rstrip("/")
if not BASE_URL:
    raise RuntimeError("REACT_APP_BACKEND_URL missing")
API = f"{BASE_URL}/api"
MONGO_URL = be.get("MONGO_URL") or os.environ.get("MONGO_URL")
DB_NAME = be.get("DB_NAME") or os.environ.get("DB_NAME")

SFX = "I115"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
STATE = {}


def _creds():
    p = Path("/app/memory/test_credentials.md")
    if not p.exists():
        pytest.skip("missing test credentials")
    return {"owner": "owner@demo.local", "ops": "ops@demo.local", "marketing": "marketing@demo.local", "pwd": "demo12345"}


@pytest.fixture(scope="module")
def creds():
    return _creds()


def _login(email, pwd):
    r = requests.post(f"{API}/auth/login", json={"email": email, "password": pwd}, timeout=60)
    if r.status_code != 200:
        pytest.fail(f"login {email} failed {r.status_code}: {r.text[:300]}")
    data = r.json()
    tok = data.get("token") or data.get("access_token")
    assert tok, f"no token in login response: {data}"
    return tok


@pytest.fixture(scope="module")
def owner(creds):
    s = requests.Session()
    s.headers.update({"Authorization": f"Bearer {_login(creds['owner'], creds['pwd'])}"})
    return s


@pytest.fixture(scope="module")
def db():
    cli = MongoClient(MONGO_URL)
    yield cli[DB_NAME]
    cli.close()


# ------------------------------------------------------------------ template & schema
class TestTemplateAndSheets:
    def test_template_download(self, owner):
        r = owner.get(f"{API}/data/import/template.xlsx", timeout=120)
        assert r.status_code == 200, r.text[:300]
        assert XLSX_MIME in r.headers.get("content-type", "")
        assert "template_master_data_rahazatrans.xlsx" in r.headers.get("content-disposition", "")
        wb = load_workbook(io.BytesIO(r.content))
        assert wb.sheetnames == ["Petunjuk", "Pelanggan", "Armada", "Driver", "Kota", "Mitra", "Add-on"], wb.sheetnames
        ws = wb["Pelanggan"]
        headers = [c.value for c in ws[1]]
        assert headers[0] == "Nama*" and "Telepon" in headers
        assert ws.cell(row=2, column=1).value == "PT Maju Jaya"  # contoh row
        assert wb["Armada"].cell(row=2, column=2).value == "D 1234 AB"
        STATE["template"] = r.content

    def test_sheets_metadata(self, owner):
        r = owner.get(f"{API}/data/import/sheets", timeout=60)
        assert r.status_code == 200
        data = r.json()
        assert len(data) == 6
        keys = [s["key"] for s in data]
        assert keys == ["customers", "vehicles", "drivers", "cities", "partners", "addons"]
        cust = next(s for s in data if s["key"] == "customers")
        assert cust["title"] == "Pelanggan"
        req = [c["field"] for c in cust["columns"] if c["required"]]
        assert req == ["name"]
        veh = next(s for s in data if s["key"] == "vehicles")
        assert sorted(c["field"] for c in veh["columns"] if c["required"]) == ["name", "plate_number"]


# ------------------------------------------------------------------ workbook builder
def _build_workbook(db, *, drop_customer_name_col=False, unknown_only=False):
    wb = load_workbook(io.BytesIO(STATE["template"]))
    if unknown_only:
        for n in list(wb.sheetnames):
            wb.remove(wb[n])
        wb.create_sheet("Sheet Aneh").cell(row=1, column=1, value="x")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    seed = db.customers.find_one({"deleted": {"$ne": True}, "phone": {"$nin": [None, ""]},
                                  "name": {"$not": {"$regex": "TEST_"}}}, {"_id": 0})
    assert seed, "no seeded customer with phone found"
    STATE["seed_customer"] = seed
    STATE["seed_new_address"] = f"Jl. Uji Impor {SFX} No. 7"

    c = wb["Pelanggan"]
    c.append([f"TEST_{SFX} Korporat Nusantara", "081955500115", f"korporat{SFX.lower()}@uji.local", "Korporat", f"Kota Uji X {SFX}", "Jl. Merdeka 11", "impor uji"])
    c.append([f"TEST_{SFX} Bapak Perorangan", "081955500116", "", "individual", "Bandung", "Jl. Kopo 5", ""])
    c.append(["", "081955500117", "", "individual", "", "Tanpa nama", ""])          # error: nama wajib
    c.append([f"TEST_{SFX} Salah Jenis", "081955500118", "", "vip", "", "", ""])      # error: choice
    c.append([seed["name"], seed["phone"], seed.get("email") or "", "", "", STATE["seed_new_address"], ""])  # update
    if drop_customer_name_col:
        c.delete_cols(1)

    v = wb["Armada"]
    v.append([f"TEST_{SFX} Hiace Uji", "F 1151 AA", "", "", 14, "tersedia", 2023, "Putih", "15/03/2027", "", 1200, ""])
    v.append([f"TEST_{SFX} Elf Uji", "F 1152 BB", "", "", 19, "", 2021, "Silver", "", "", "", ""])

    wb["Driver"].append([f"TEST_{SFX} Driver Uji", "081955500120", "9999-0001-0002", "2029-01-31", "", 275000])
    wb["Kota"].append([f"Kota Uji Y {SFX}"])
    wb["Mitra"].append([f"TEST_{SFX} Mitra Uji", "Pak Uji", "081955500121", "", "Semarang", "Jl. Uji 1", "active", ""])
    a = wb["Add-on"]
    a.append([f"TEST_{SFX} Addon Aktif", 150000, "ya"])
    a.append([f"TEST_{SFX} Addon Nonaktif", 50000, "tidak"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _f(data, name="import_uji_i115.xlsx", mime=XLSX_MIME):
    return {"file": (name, data, mime)}


COLS = ["customers", "vehicles", "drivers", "cities", "partners", "addons"]


def _counts(db):
    return {k: db[k].count_documents({}) for k in COLS}


# ------------------------------------------------------------------ preview
class TestPreview:
    def test_preview_counts_and_no_write(self, owner, db):
        STATE["wb"] = _build_workbook(db)
        before = _counts(db)
        r = owner.post(f"{API}/data/import/preview", files=_f(STATE["wb"]), timeout=180)
        assert r.status_code == 200, r.text[:500]
        sh = r.json()["sheets"]
        assert set(sh) == set(COLS)
        cu = sh["customers"]
        assert (cu["total"], cu["valid"], cu["insert"], cu["update"]) == (5, 3, 2, 1), cu
        assert len(cu["errors"]) == 2
        msgs = " | ".join(f"Baris {e['row']}: {e['msg']}" for e in cu["errors"])
        assert "wajib diisi" in msgs and "vip" in msgs, msgs
        assert all(isinstance(e["row"], int) and e["row"] >= 3 for e in cu["errors"])
        # example rows must be skipped everywhere
        assert (sh["vehicles"]["total"], sh["vehicles"]["valid"], sh["vehicles"]["insert"]) == (2, 2, 2), sh["vehicles"]
        assert (sh["drivers"]["total"], sh["drivers"]["insert"]) == (1, 1), sh["drivers"]
        assert (sh["cities"]["total"], sh["cities"]["insert"]) == (1, 1), sh["cities"]
        assert (sh["partners"]["total"], sh["partners"]["insert"]) == (1, 1), sh["partners"]
        assert (sh["addons"]["total"], sh["addons"]["insert"]) == (2, 2), sh["addons"]
        for k, s in sh.items():
            assert s["missing_columns"] == [], (k, s["missing_columns"])
        assert _counts(db) == before, "preview must not write to DB"

    def test_preview_rejects_non_xlsx(self, owner):
        for name, mime in [("dump.zip", "application/zip"), ("notes.txt", "text/plain")]:
            r = owner.post(f"{API}/data/import/preview", files=_f(b"junk", name, mime), timeout=60)
            assert r.status_code == 400, (name, r.status_code, r.text[:200])
            assert "Excel" in r.json()["detail"]

    def test_preview_unknown_sheets_400(self, owner, db):
        r = owner.post(f"{API}/data/import/preview", files=_f(_build_workbook(db, unknown_only=True)), timeout=60)
        assert r.status_code == 400
        assert "sheet" in r.json()["detail"].lower()

    def test_preview_missing_required_column(self, owner, db):
        data = _build_workbook(db, drop_customer_name_col=True)
        r = owner.post(f"{API}/data/import/preview", files=_f(data), timeout=180)
        assert r.status_code == 200, r.text[:300]
        cu = r.json()["sheets"]["customers"]
        assert cu["missing_columns"] == ["Nama*"], cu
        assert cu["valid"] == 0 and cu["total"] == 0
        STATE["wb_missing_col"] = data


# ------------------------------------------------------------------ commit
class TestCommitUpsert:
    def test_commit_upsert_with_snapshot(self, owner, db):
        before = _counts(db)
        STATE["before_counts"] = before
        r = owner.post(f"{API}/data/import/commit", files=_f(STATE["wb"]),
                       data={"mode": "upsert", "sheets": "", "snapshot": "true"}, timeout=300)
        assert r.status_code == 200, r.text[:500]
        body = r.json()
        assert "_id" not in body
        tot = body["totals"]
        assert tot == {"inserted": 9, "updated": 1, "skipped": 0, "errors": 2}, tot
        assert body["mode"] == "upsert"
        assert body["snapshot_id"], "snapshot_id must be set"
        STATE["snapshot_id"] = body["snapshot_id"]
        s = body["summary"]
        assert (s["customers"]["inserted"], s["customers"]["updated"], s["customers"]["errors"]) == (2, 1, 2)
        assert s["vehicles"]["inserted"] == 2 and s["addons"]["inserted"] == 2
        after = _counts(db)
        assert after["customers"] == before["customers"] + 2
        assert after["vehicles"] == before["vehicles"] + 2
        # cities: 1 explicit + 'Kota Uji X I115' auto-created (Bandung already exists)
        assert after["cities"] >= before["cities"] + 2, (before["cities"], after["cities"])

    def test_snapshot_archive_in_overview(self, owner):
        r = owner.get(f"{API}/data/overview", timeout=180)
        assert r.status_code == 200
        backups = r.json()["backups"]
        hit = next((b for b in backups if b["id"] == STATE["snapshot_id"]), None)
        assert hit, "snapshot archive missing from overview"
        assert hit["kind"] == "pre_import", hit
        assert "impor" in (hit.get("note") or "").lower()

    def test_inserted_docs_shape(self, db):
        c = db.customers.find_one({"name": f"TEST_{SFX} Korporat Nusantara"}, {"_id": 0})
        assert c, "new customer not persisted"
        assert c["type"] == "corporate", c["type"]
        assert c["city"] == f"Kota Uji X {SFX}"
        assert c["phone_normalized"] and c["phone_normalized"].endswith("955500115"), c["phone_normalized"]
        assert c["id"].startswith("cus") and c["total_trips"] == 0
        assert db.cities.find_one({"name": f"Kota Uji X {SFX}"}), "auto city not created"
        assert db.cities.find_one({"name": f"Kota Uji Y {SFX}"}), "sheet city not created"

        v = db.vehicles.find_one({"plate_number": "F 1151 AA"}, {"_id": 0})
        assert v, "vehicle not persisted"
        assert v["status"] == "available", v["status"]
        assert v["kir_expiry"] == "2027-03-15", v["kir_expiry"]
        assert re.fullmatch(r"V-\d{2}", v.get("code") or ""), v.get("code")
        v2 = db.vehicles.find_one({"plate_number": "F 1152 BB"}, {"_id": 0})
        assert v2 and v2["code"] != v["code"], "auto codes must be unique"

        d = db.drivers.find_one({"name": f"TEST_{SFX} Driver Uji"}, {"_id": 0})
        assert d and d["sim_expiry"] == "2029-01-31" and d["default_fee_rate"] == 275000, d
        assert db.partners.find_one({"name": f"TEST_{SFX} Mitra Uji"})
        a1 = db.addons.find_one({"label": f"TEST_{SFX} Addon Aktif"}, {"_id": 0})
        a2 = db.addons.find_one({"label": f"TEST_{SFX} Addon Nonaktif"}, {"_id": 0})
        assert a1 and a1["active"] is True and a1["default_amount"] == 150000, a1
        assert a2 and a2["active"] is False, a2

    def test_existing_customer_updated(self, db):
        seed = STATE["seed_customer"]
        cur = db.customers.find_one({"id": seed["id"]}, {"_id": 0})
        assert cur, "seed customer disappeared"
        assert cur["address"] == STATE["seed_new_address"], cur["address"]
        assert cur.get("updated_at")
        assert db.customers.count_documents({"name": seed["name"], "deleted": {"$ne": True}}) == 1

    def test_recommit_same_file_is_idempotent(self, owner, db):
        before = _counts(db)
        r = owner.post(f"{API}/data/import/commit", files=_f(STATE["wb"]),
                       data={"mode": "upsert", "snapshot": "false"}, timeout=300)
        assert r.status_code == 200, r.text[:400]
        tot = r.json()["totals"]
        assert tot["inserted"] == 0, tot
        assert tot["updated"] == 10, tot
        assert r.json()["snapshot_id"] is None
        assert _counts(db) == before, "re-commit must not create duplicates"

    def test_insert_only_skips_matches(self, owner, db):
        before = _counts(db)
        r = owner.post(f"{API}/data/import/commit", files=_f(STATE["wb"]),
                       data={"mode": "insert_only", "snapshot": "false"}, timeout=300)
        assert r.status_code == 200, r.text[:400]
        tot = r.json()["totals"]
        assert (tot["inserted"], tot["updated"], tot["skipped"]) == (0, 0, 10), tot
        assert _counts(db) == before

    def test_sheets_filter_only_processes_selected(self, owner, db):
        r = owner.post(f"{API}/data/import/commit", files=_f(STATE["wb"]),
                       data={"mode": "upsert", "sheets": "cities", "snapshot": "false"}, timeout=300)
        assert r.status_code == 200, r.text[:400]
        body = r.json()
        assert list(body["summary"].keys()) == ["cities"], body["summary"]
        assert body["summary"]["cities"]["updated"] == 1

    def test_commit_missing_required_column_sheet_skipped(self, owner, db):
        before = db.customers.count_documents({})
        r = owner.post(f"{API}/data/import/commit", files=_f(STATE["wb_missing_col"]),
                       data={"mode": "upsert", "snapshot": "false"}, timeout=300)
        assert r.status_code == 200, r.text[:400]
        cu = r.json()["summary"]["customers"]
        assert cu["inserted"] == 0 and cu["updated"] == 0, cu
        assert cu.get("note") == "kolom wajib hilang", cu
        assert db.customers.count_documents({}) == before

    def test_commit_invalid_mode_400(self, owner):
        r = owner.post(f"{API}/data/import/commit", files=_f(STATE["wb"]),
                       data={"mode": "replace_all", "snapshot": "false"}, timeout=120)
        assert r.status_code == 400, r.text[:200]
        assert "Mode" in r.json()["detail"]

    def test_commit_rejects_non_xlsx(self, owner):
        r = owner.post(f"{API}/data/import/commit", files=_f(b"junk", "x.zip", "application/zip"),
                       data={"mode": "upsert", "snapshot": "false"}, timeout=120)
        assert r.status_code == 400

    def test_commit_unknown_sheet_workbook_400(self, owner, db):
        r = owner.post(f"{API}/data/import/commit", files=_f(_build_workbook(db, unknown_only=True)),
                       data={"mode": "upsert", "snapshot": "false"}, timeout=120)
        assert r.status_code == 400


# ------------------------------------------------------------------ history & RBAC
class TestHistoryAndRbac:
    def test_history(self, owner):
        r = owner.get(f"{API}/data/import/history", timeout=60)
        assert r.status_code == 200
        rows = r.json()
        assert isinstance(rows, list) and rows
        assert all("_id" not in x for x in rows)
        latest = rows[0]
        assert {"id", "mode", "summary", "finished_at", "actor_name"} <= set(latest)
        assert rows[0]["finished_at"] >= rows[-1]["finished_at"]
        assert any(x["mode"] == "insert_only" for x in rows)

    def test_ops_admin_allowed(self, creds):
        s = requests.Session()
        s.headers.update({"Authorization": f"Bearer {_login(creds['ops'], creds['pwd'])}"})
        for path in ("/data/import/sheets", "/data/import/history"):
            r = s.get(f"{API}{path}", timeout=60)
            assert r.status_code == 200, (path, r.status_code)

    def test_marketing_forbidden(self, creds):
        s = requests.Session()
        s.headers.update({"Authorization": f"Bearer {_login(creds['marketing'], creds['pwd'])}"})
        for path in ("/data/import/sheets", "/data/import/history", "/data/import/template.xlsx"):
            r = s.get(f"{API}{path}", timeout=60)
            assert r.status_code == 403, (path, r.status_code)
        r = s.post(f"{API}/data/import/preview", files=_f(STATE["wb"]), timeout=120)
        assert r.status_code == 403

    def test_unauthenticated_401(self):
        r = requests.get(f"{API}/data/import/sheets", timeout=60)
        assert r.status_code in (401, 403), r.status_code


# ------------------------------------------------------------------ cleanup
class TestCleanup:
    def test_cleanup_test_data(self, owner, db):
        db.customers.delete_many({"name": {"$regex": f"^TEST_{SFX}"}})
        db.vehicles.delete_many({"name": {"$regex": f"^TEST_{SFX}"}})
        db.drivers.delete_many({"name": {"$regex": f"^TEST_{SFX}"}})
        db.partners.delete_many({"name": {"$regex": f"^TEST_{SFX}"}})
        db.addons.delete_many({"label": {"$regex": f"^TEST_{SFX}"}})
        db.cities.delete_many({"name": {"$regex": f"^Kota Uji [XY] {SFX}$"}})
        seed = STATE.get("seed_customer")
        if seed:
            db.customers.update_one({"id": seed["id"]}, {"$set": {"address": seed.get("address", "")}})
        # drop extra pre_import archives created by this suite
        r = owner.get(f"{API}/data/overview", timeout=180)
        for b in [x for x in r.json()["backups"] if x["kind"] == "pre_import"]:
            owner.delete(f"{API}/data/backups/{b['id']}", timeout=120)
        assert db.customers.count_documents({"name": {"$regex": f"^TEST_{SFX}"}}) == 0
