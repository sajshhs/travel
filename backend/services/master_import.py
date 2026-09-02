"""services/master_import.py — impor MASTER DATA massal dari satu workbook Excel (migrasi):
Pelanggan, Armada, Driver, Kota, Mitra, Add-on — semua sheet sekaligus, mode upsert / hanya-tambah.

Alur: template.xlsx (header + contoh) → preview (validasi, hitung tambah/perbarui/error, tanpa menulis)
→ commit (snapshot backup `pre_import` dulu, lalu tulis). Kunci pencocokan per sheet (telepon/plat/nama).
"""
import re
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core_utils import money, new_id, now_iso
from services.identity import normalize_phone
from services.pricing import VEHICLE_TYPES
from services.refs import CUSTOMER_TYPES, DRIVER_STATUSES, VEHICLE_STATUSES

ALIAS = {
    "individu": "individual", "perorangan": "individual", "korporat": "corporate", "perusahaan": "corporate",
    "tersedia": "available", "perawatan": "maintenance", "servis": "maintenance", "nonaktif": "inactive", "terjual": "sold", "jalan": "on_trip",
    "istirahat": "resting", "aktif": "active", "ya": True, "tidak": False, "y": True, "n": False, "true": True, "false": False,
}
# (header, field, required, tipe, pilihan) — tipe: str|phone|email|int|float|money|date|choice|bool
SHEETS = {
    "customers": {"title": "Pelanggan", "collection": "customers", "id_prefix": "cus", "columns": [
        ("Nama*", "name", True, "str", None), ("Telepon", "phone", False, "phone", None), ("Email", "email", False, "email", None),
        ("Jenis (individual/corporate)", "type", False, "choice", CUSTOMER_TYPES), ("Kota", "city", False, "city", None),
        ("Alamat", "address", False, "str", None), ("Catatan", "notes", False, "str", None)],
        "example": ["PT Maju Jaya", "081234567890", "admin@majujaya.co.id", "corporate", "Bandung", "Jl. Asia Afrika 1", "Langganan tahunan"]},
    "vehicles": {"title": "Armada", "collection": "vehicles", "id_prefix": "veh", "columns": [
        ("Nama*", "name", True, "str", None), ("Plat Nomor*", "plate_number", True, "plate", None), ("Kode Unit", "code", False, "str", None),
        ("Tipe (" + "/".join(VEHICLE_TYPES) + ")", "type", False, "choice", VEHICLE_TYPES), ("Kapasitas", "capacity", False, "int", None),
        ("Status (" + "/".join(VEHICLE_STATUSES) + ")", "status", False, "choice", VEHICLE_STATUSES), ("Tahun", "year", False, "int", None),
        ("Warna", "color", False, "str", None), ("KIR Berlaku s/d", "kir_expiry", False, "date", None), ("Pajak Berlaku s/d", "tax_expiry", False, "date", None),
        ("Odometer (km)", "odometer", False, "float", None), ("Catatan", "notes", False, "str", None)],
        "example": ["Hiace Premio 01", "D 1234 AB", "V-01", "hiace_premio", 14, "available", 2022, "Putih", "2027-01-31", "2026-12-31", 45000, ""]},
    "drivers": {"title": "Driver", "collection": "drivers", "id_prefix": "drv", "columns": [
        ("Nama*", "name", True, "str", None), ("Telepon", "phone", False, "phone", None), ("No. SIM", "sim_number", False, "str", None),
        ("SIM Berlaku s/d", "sim_expiry", False, "date", None), ("Status (" + "/".join(DRIVER_STATUSES) + ")", "status", False, "choice", DRIVER_STATUSES),
        ("Fee Default per Hari (Rp)", "default_fee_rate", False, "money", None)],
        "example": ["Budi Santoso", "081298765432", "1234-5678-9012", "2028-06-30", "offline", 250000]},
    "cities": {"title": "Kota", "collection": "cities", "id_prefix": "cty", "columns": [("Nama Kota*", "name", True, "str", None)],
               "example": ["Bandung"]},
    "partners": {"title": "Mitra", "collection": "partners", "id_prefix": "ptn", "columns": [
        ("Nama Mitra*", "name", True, "str", None), ("PIC", "pic", False, "str", None), ("Telepon", "phone", False, "phone", None),
        ("Email", "email", False, "email", None), ("Kota", "city", False, "city", None), ("Alamat", "address", False, "str", None),
        ("Status (active/inactive)", "status", False, "choice", ("active", "inactive")), ("Catatan", "notes", False, "str", None)],
        "example": ["CV Armada Sejahtera", "Pak Dedi", "082111222333", "dedi@armada.id", "Jakarta", "Jl. Sudirman 10", "active", "Mitra unit bus"]},
    "addons": {"title": "Add-on", "collection": "addons", "id_prefix": "add", "columns": [
        ("Label*", "label", True, "str", None), ("Nominal Default (Rp)", "default_amount", False, "money", None), ("Aktif (ya/tidak)", "active", False, "bool", None)],
        "example": ["Overtime per jam", 100000, "ya"]},
}
TITLE_TO_KEY = {v["title"].lower(): k for k, v in SHEETS.items()}


def _norm_header(h) -> str:
    return re.sub(r"[^a-z]", "", str(h or "").split("(")[0].lower())


def _parse_date(v):
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"tanggal '{s}' tidak dikenali (pakai YYYY-MM-DD atau DD/MM/YYYY)")


def _coerce(kind, v, choices):
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if kind in ("str", "city"):
        return str(v).strip()
    if kind == "phone":
        s = str(v).strip()
        if s.endswith(".0"):
            s = s[:-2]
        return s
    if kind == "email":
        s = str(v).strip().lower()
        if "@" not in s:
            raise ValueError(f"email '{s}' tidak valid")
        return s
    if kind == "plate":
        return re.sub(r"\s+", " ", str(v).strip().upper())
    if kind == "int":
        return int(float(v))
    if kind in ("float", "money"):
        s = re.sub(r"[^\d,.\-]", "", str(v)) if isinstance(v, str) else v
        return money(float(str(s).replace(".", "").replace(",", ".")) if isinstance(s, str) and s.count(",") == 1 and s.count(".") > 0 else float(s)) if kind == "money" else float(s)
    if kind == "date":
        return _parse_date(v)
    if kind == "bool":
        if isinstance(v, bool):
            return v
        a = ALIAS.get(str(v).strip().lower())
        if isinstance(a, bool):
            return a
        raise ValueError(f"nilai '{v}' harus ya/tidak")
    if kind == "choice":
        s = str(v).strip().lower().replace(" ", "_")
        s = ALIAS.get(s, s) if isinstance(ALIAS.get(s, s), str) else s
        if s not in choices:
            raise ValueError(f"nilai '{v}' tidak ada di pilihan {'/'.join(choices)}")
        return s
    return v


def build_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Petunjuk"
    for i, line in enumerate([
        "TEMPLATE IMPOR MASTER DATA — RahazaTrans",
        "1. Isi sheet yang diperlukan (Pelanggan, Armada, Driver, Kota, Mitra, Add-on). Sheet kosong dilewati.",
        "2. Kolom bertanda * wajib. Jangan mengubah nama kolom di baris 1. Baris 2 adalah CONTOH — hapus/ganti.",
        "3. Tanggal: YYYY-MM-DD atau DD/MM/YYYY. Nominal: angka tanpa 'Rp'.",
        "4. Pencocokan data lama (mode perbarui): Pelanggan → telepon/email/nama; Armada → plat; Driver → telepon/nama; Kota/Mitra/Add-on → nama.",
        "5. Kota pada Pelanggan/Mitra yang belum ada di master Kota akan dibuat otomatis.",
        "6. Unggah di Manajemen Data → Impor Master Data → Pratinjau → Impor Semua Sekaligus.",
    ], 1):
        c = ws.cell(row=i, column=1, value=line)
        c.font = Font(bold=(i == 1), size=13 if i == 1 else 11)
    ws.column_dimensions["A"].width = 120
    head_font, head_fill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="0F6E56")
    for spec in SHEETS.values():
        s = wb.create_sheet(spec["title"])
        for ci, (header, *_rest) in enumerate(spec["columns"], 1):
            c = s.cell(row=1, column=ci, value=header)
            c.font, c.fill, c.alignment = head_font, head_fill, Alignment(vertical="center")
            s.column_dimensions[get_column_letter(ci)].width = max(14, min(len(header) + 4, 40))
        for ci, v in enumerate(spec["example"], 1):
            s.cell(row=2, column=ci, value=v).font = Font(italic=True, color="8E8E93")
        s.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _is_example_row(spec, raw, colidx) -> bool:
    """Baris contoh template yang tidak dihapus pengguna → dilewati diam-diam."""
    vals = []
    for (_, field, *_r), ex in zip(spec["columns"], spec["example"]):
        idx = colidx.get(field)
        v = raw[idx] if idx is not None and idx < len(raw) else None
        v_s, ex_s = ("" if v is None else str(v).strip()), ("" if ex is None else str(ex).strip())
        vals.append(v_s == ex_s)
    return all(vals)


def parse_workbook(data: bytes) -> dict:
    """{sheet_key: {"rows": [{field: value, "_row": n}], "errors": [{"row", "msg"}], "missing_columns": []}}"""
    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"Berkas Excel tidak terbaca ({exc}).")
    out = {}
    for ws in wb.worksheets:
        key = TITLE_TO_KEY.get(ws.title.strip().lower())
        if not key:
            continue
        spec = SHEETS[key]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            continue
        hmap = {_norm_header(h): i for i, h in enumerate(header) if h is not None}
        colidx = {}
        missing = []
        for hdr, field, req, *_ in spec["columns"]:
            idx = hmap.get(_norm_header(hdr))
            if idx is None and req:
                missing.append(hdr)
            colidx[field] = idx
        res = {"rows": [], "errors": [], "missing_columns": missing}
        if missing:
            out[key] = res
            continue
        for rn, raw in enumerate(rows_iter, 2):
            if raw is None or all(v in (None, "") for v in raw):
                continue
            if rn == 2 and _is_example_row(spec, raw, colidx):
                continue
            rec, errs = {"_row": rn}, []
            for hdr, field, req, kind, choices in spec["columns"]:
                idx = colidx.get(field)
                v = raw[idx] if idx is not None and idx < len(raw) else None
                try:
                    val = _coerce(kind, v, choices)
                except (ValueError, TypeError) as exc:
                    errs.append(f"{hdr.split('(')[0].strip().rstrip('*')}: {exc}")
                    val = None
                if req and val in (None, ""):
                    errs.append(f"{hdr.rstrip('*')} wajib diisi")
                rec[field] = val
            if errs:
                res["errors"].append({"row": rn, "msg": "; ".join(errs)})
            else:
                res["rows"].append(rec)
        out[key] = res
    if not out:
        raise ValueError("Tidak ada sheet yang dikenali. Pakai template: sheet Pelanggan/Armada/Driver/Kota/Mitra/Add-on.")
    return out


# ------------------------------------------------------------------ pencocokan & penulisan
def _ci(v):
    return re.sub(r"\s+", " ", str(v or "").strip().lower())


async def _existing_index(db, key: str) -> dict:
    """Peta kunci → dokumen lama untuk pencocokan cepat."""
    col = SHEETS[key]["collection"]
    idx = {}
    async for d in db[col].find({"deleted": {"$ne": True}}, {"_id": 0}):
        if key == "customers":
            for k in (("p:" + normalize_phone(d.get("phone"))) if d.get("phone") else None, ("e:" + _ci(d.get("email"))) if d.get("email") else None, "n:" + _ci(d.get("name"))):
                if k:
                    idx.setdefault(k, d)
        elif key == "vehicles":
            idx.setdefault("pl:" + re.sub(r"\s", "", _ci(d.get("plate_number"))), d)
        elif key == "drivers":
            if d.get("phone"):
                idx.setdefault("p:" + normalize_phone(d.get("phone")), d)
            idx.setdefault("n:" + _ci(d.get("name")), d)
        else:
            idx.setdefault("n:" + _ci(d.get("name") or d.get("label")), d)
    return idx


def _match(key: str, rec: dict, idx: dict):
    if key == "customers":
        for k in (("p:" + normalize_phone(rec.get("phone"))) if rec.get("phone") else None, ("e:" + _ci(rec.get("email"))) if rec.get("email") else None, "n:" + _ci(rec.get("name"))):
            if k and k in idx:
                return idx[k]
        return None
    if key == "vehicles":
        return idx.get("pl:" + re.sub(r"\s", "", _ci(rec.get("plate_number"))))
    if key == "drivers":
        if rec.get("phone") and ("p:" + normalize_phone(rec.get("phone"))) in idx:
            return idx["p:" + normalize_phone(rec.get("phone"))]
        return idx.get("n:" + _ci(rec.get("name")))
    return idx.get("n:" + _ci(rec.get("name") or rec.get("label")))


def _new_doc(key: str, rec: dict) -> dict:
    ts = now_iso()
    base = {k: v for k, v in rec.items() if not k.startswith("_") and v is not None}
    if key == "customers":
        return {"id": new_id("cus"), "name": "", "phone": "", "email": "", "type": "individual", "city": "", "address": "", "notes": "",
                "total_trips": 0, "lifetime_value": 0.0, "created_at": ts, **base, "phone_normalized": normalize_phone(rec.get("phone"))}
    if key == "vehicles":
        return {"id": new_id("veh"), "type": "hiace", "capacity": 0, "status": "available", "odometer": 0.0, "features": [], "photos": [], "gallery": [],
                "tour_scenes": [], "specs": [], "highlights": [], "price_from": None, "day_rate": 0, "publish_to_web": True, "ownership": "own",
                "partner_id": None, "notes": "", "created_at": ts, **base}
    if key == "drivers":
        return {"id": new_id("drv"), "phone": "", "sim_number": "", "sim_expiry": None, "status": "offline", "current_vehicle_id": None, "rating": 0.0,
                "default_fee_rate": None, "created_at": ts, **base}
    if key == "cities":
        return {"id": new_id("cty"), "active": True, "created_at": ts, **base}
    if key == "partners":
        return {"id": new_id("ptn"), "pic": "", "phone": "", "email": "", "city": "", "address": "", "rating": 0.0, "notes": "", "status": "active", "created_at": ts, **base}
    return {"id": new_id("add"), "default_amount": 0, "active": True, "created_at": ts, **base}


async def _ensure_city(db, name: str, cache: dict) -> str:
    if not name:
        return ""
    low = _ci(name)
    if low in cache:
        return cache[low]
    hit = await db.cities.find_one({"deleted": {"$ne": True}, "name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}}, {"_id": 0})
    if hit:
        if hit.get("active") is False:
            await db.cities.update_one({"id": hit["id"]}, {"$set": {"active": True}})
        cache[low] = hit["name"]
        return hit["name"]
    await db.cities.insert_one({"id": new_id("cty"), "name": name, "active": True, "created_at": now_iso(), "source": "import"})
    cache[low] = name
    return name


async def plan(db, parsed: dict) -> dict:
    """Pratinjau tanpa menulis: hitung tambah/perbarui per sheet + duplikat dalam berkas."""
    out = {}
    for key, res in parsed.items():
        idx = await _existing_index(db, key)
        seen, insert, update, dup = set(), 0, 0, []
        for rec in res["rows"]:
            sig = (re.sub(r"\s", "", _ci(rec.get("plate_number"))) if key == "vehicles" else _ci(rec.get("name") or rec.get("label")))
            if sig in seen:
                dup.append({"row": rec["_row"], "msg": "duplikat di dalam berkas (baris ini akan menimpa baris sebelumnya)"})
            seen.add(sig)
            if _match(key, rec, idx):
                update += 1
            else:
                insert += 1
        out[key] = {"title": SHEETS[key]["title"], "total": len(res["rows"]) + len(res["errors"]), "valid": len(res["rows"]), "insert": insert,
                    "update": update, "errors": res["errors"], "warnings": dup, "missing_columns": res["missing_columns"],
                    "sample": [{k: v for k, v in r.items() if not k.startswith("_")} for r in res["rows"][:3]]}
    return out


async def commit(db, parsed: dict, *, mode: str = "upsert", sheets=None, actor: dict) -> dict:
    """mode: upsert (tambah+perbarui) | insert_only (yang cocok dilewati). Baris error selalu dilewati."""
    summary = {}
    city_cache = {}
    for key, res in parsed.items():
        if sheets and key not in sheets:
            continue
        if res["missing_columns"]:
            summary[key] = {"title": SHEETS[key]["title"], "inserted": 0, "updated": 0, "skipped": len(res["rows"]), "errors": len(res["errors"]), "note": "kolom wajib hilang"}
            continue
        col = SHEETS[key]["collection"]
        idx = await _existing_index(db, key)
        ins = upd = skip = 0
        for rec in res["rows"]:
            if key in ("customers", "partners") and rec.get("city"):
                rec["city"] = await _ensure_city(db, rec["city"], city_cache)
            old = _match(key, rec, idx)
            if old:
                if mode == "insert_only":
                    skip += 1
                    continue
                patch = {k: v for k, v in rec.items() if not k.startswith("_") and v is not None}
                if key == "customers":
                    patch["phone_normalized"] = normalize_phone(patch.get("phone", old.get("phone")))
                patch["updated_at"] = now_iso()
                await db[col].update_one({"id": old["id"]}, {"$set": patch})
                upd += 1
            else:
                doc = _new_doc(key, rec)
                if key == "vehicles" and not doc.get("code"):
                    codes = set(await db.vehicles.distinct("code"))
                    n = 1
                    while f"V-{n:02d}" in codes:
                        n += 1
                    doc["code"] = f"V-{n:02d}"
                await db[col].insert_one(dict(doc))
                doc.pop("_id", None)
                # daftarkan ke indeks agar baris duplikat berikutnya jadi update
                idx.update({k: doc for k in _keys_for(key, doc)})
                ins += 1
        summary[key] = {"title": SHEETS[key]["title"], "inserted": ins, "updated": upd, "skipped": skip, "errors": len(res["errors"])}
    log = {"id": new_id("imp"), "mode": mode, "summary": summary, "actor_id": actor.get("id"), "actor_name": actor.get("name"), "finished_at": now_iso()}
    await db.import_logs.insert_one(dict(log))
    log.pop("_id", None)
    return log


def _keys_for(key: str, d: dict) -> list:
    if key == "customers":
        return [k for k in (("p:" + normalize_phone(d.get("phone"))) if d.get("phone") else None, ("e:" + _ci(d.get("email"))) if d.get("email") else None, "n:" + _ci(d.get("name"))) if k]
    if key == "vehicles":
        return ["pl:" + re.sub(r"\s", "", _ci(d.get("plate_number")))]
    if key == "drivers":
        return [k for k in (("p:" + normalize_phone(d.get("phone"))) if d.get("phone") else None, "n:" + _ci(d.get("name"))) if k]
    return ["n:" + _ci(d.get("name") or d.get("label"))]
