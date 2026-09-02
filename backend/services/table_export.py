"""services/table_export.py — ekspor koleksi bisnis ke Excel (satu sheet per tabel, kolom ramah manusia)."""
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# kolom: (key, header). Nilai bersarang dipipihkan oleh _cell().
TABLES = {
    "customers": {"label": "Pelanggan", "collection": "customers", "sort": "name", "columns": [
        ("name", "Nama"), ("phone", "Telepon"), ("email", "Email"), ("company", "Perusahaan"), ("address", "Alamat"),
        ("segment", "Segmen"), ("tags", "Tag"), ("total_bookings", "Jumlah Booking"), ("total_spent", "Total Belanja (Rp)"),
        ("notes", "Catatan"), ("created_at", "Dibuat")]},
    "bookings": {"label": "Booking", "collection": "bookings", "sort": "-start_datetime", "columns": [
        ("code", "Kode"), ("status", "Status"), ("customer_name", "Pelanggan"), ("customer_phone", "Telepon"),
        ("vehicle_name", "Armada"), ("driver_name", "Driver"), ("origin", "Titik Jemput"), ("destination", "Tujuan"),
        ("start_datetime", "Mulai"), ("end_datetime", "Selesai"), ("pax", "Penumpang"), ("base_price", "Harga Dasar (Rp)"),
        ("total_amount", "Total (Rp)"), ("paid_amount", "Terbayar (Rp)"), ("dp_amount", "DP (Rp)"), ("source", "Sumber"),
        ("notes", "Catatan"), ("cancellation_reason", "Alasan Batal"), ("refund_amount", "Refund (Rp)"), ("created_at", "Dibuat")]},
    "payments": {"label": "Pembayaran", "collection": "payments", "sort": "-paid_at", "columns": [
        ("paid_at", "Tanggal Bayar"), ("booking_code", "Kode Booking"), ("customer_name", "Pelanggan"), ("type", "Jenis"),
        ("method", "Metode"), ("amount", "Nominal (Rp)"), ("reference", "Referensi"), ("notes", "Catatan"),
        ("recorded_by_name", "Dicatat Oleh"), ("created_at", "Dibuat")]},
    "invoices": {"label": "Invoice", "collection": "invoices", "sort": "-issued_at", "columns": [
        ("number", "Nomor"), ("kind_label", "Jenis"), ("status", "Status"), ("booking_code", "Kode Booking"), ("customer_name", "Pelanggan"),
        ("subtotal", "Subtotal (Rp)"), ("tax_label", "Pajak"), ("tax_percent", "Pajak %"), ("tax_amount", "Nilai Pajak (Rp)"),
        ("grand_total", "Total Booking (Rp)"), ("amount_due", "Ditagihkan (Rp)"), ("issued_at", "Terbit"), ("due_at", "Jatuh Tempo"),
        ("sent_count", "Terkirim WA"), ("notes", "Catatan")]},
    "receipts": {"label": "Kwitansi", "collection": "receipts", "sort": "-issued_at", "columns": [
        ("number", "Nomor"), ("booking_code", "Kode Booking"), ("customer_name", "Pelanggan"), ("payment_type", "Jenis Bayar"),
        ("method", "Metode"), ("amount", "Nominal (Rp)"), ("amount_words", "Terbilang"), ("paid_at", "Tanggal Bayar"),
        ("remaining_after", "Sisa Tagihan (Rp)"), ("issued_at", "Terbit"), ("sent_count", "Terkirim WA")]},
    "vehicles": {"label": "Armada", "collection": "vehicles", "sort": "name", "columns": [
        ("name", "Nama"), ("plate", "Plat"), ("plate_number", "Plat (alt)"), ("type", "Tipe"), ("brand", "Merek"), ("model", "Model"),
        ("year", "Tahun"), ("capacity", "Kapasitas"), ("status", "Status"), ("odometer", "Odometer"), ("notes", "Catatan")]},
    "drivers": {"label": "Driver", "collection": "drivers", "sort": "name", "columns": [
        ("name", "Nama"), ("phone", "Telepon"), ("license_number", "No. SIM"), ("license_type", "Jenis SIM"),
        ("license_expiry", "SIM Berlaku s/d"), ("status", "Status"), ("address", "Alamat"), ("notes", "Catatan")]},
}
MONEY_KEYS = {"total_spent", "base_price", "total_amount", "paid_amount", "dp_amount", "refund_amount", "amount", "subtotal",
              "tax_amount", "grand_total", "amount_due", "remaining_after"}
DATE_KEYS = {"created_at", "start_datetime", "end_datetime", "paid_at", "issued_at", "due_at", "license_expiry"}

# Filter per tabel: field tanggal + pilihan status (key, label, opsi)
FILTERS = {
    "bookings": {"date_field": "start_datetime", "date_label": "Jadwal mulai",
                 "status": [("status", "Status", ["hold", "confirmed", "ongoing", "completed", "cancelled"])]},
    "payments": {"date_field": "paid_at", "date_label": "Tanggal bayar",
                 "status": [("type", "Jenis", ["dp", "settlement", "refund"]), ("method", "Metode", ["transfer", "cash", "qris"])]},
    "invoices": {"date_field": "issued_at", "date_label": "Tanggal terbit",
                 "status": [("status", "Status", ["draft", "sent", "partial", "paid", "void"]), ("kind", "Jenis", ["dp", "settlement", "full"])]},
    "receipts": {"date_field": "issued_at", "date_label": "Tanggal terbit", "status": [("payment_type", "Jenis bayar", ["dp", "settlement"])]},
}


def build_query(key: str, date_from: str = None, date_to: str = None, filters: dict = None) -> dict:
    spec = FILTERS.get(key)
    q = {}
    if not spec:
        return q
    if date_from or date_to:
        rng = {}
        if date_from:
            rng["$gte"] = f"{date_from[:10]}T00:00:00"
        if date_to:
            rng["$lte"] = f"{date_to[:10]}T23:59:59.999999+00:00"
        q[spec["date_field"]] = rng
    for fk, _, allowed in spec["status"]:
        vals = [v for v in (filters or {}).get(fk, []) if v in allowed]
        if vals:
            q[fk] = {"$in": vals}
    return q


def _cell(key, val):
    if val is None or val == "":
        return None
    if isinstance(val, list):
        return ", ".join(str(v.get("label", v) if isinstance(v, dict) else v) for v in val)
    if isinstance(val, dict):
        return ", ".join(f"{k}: {v}" for k, v in val.items())
    if key in DATE_KEYS and isinstance(val, str):
        try:
            d = datetime.fromisoformat(val.replace("Z", "+00:00"))
            return d.replace(tzinfo=None)
        except ValueError:
            return val
    if key in MONEY_KEYS:
        try:
            return float(val)
        except (TypeError, ValueError):
            return val
    return val


async def export_table(db, key: str, query: dict = None) -> tuple:
    spec = TABLES[key]
    sort_key = spec["sort"]
    direction = -1 if sort_key.startswith("-") else 1
    rows = await db[spec["collection"]].find(query or {}, {"_id": 0}).sort(sort_key.lstrip("-"), direction).to_list(50000)
    if key == "payments":  # pembayaran hanya simpan booking_id → lengkapi kode & nama pelanggan
        ids = list({r.get("booking_id") for r in rows if r.get("booking_id")})
        bmap = {b["id"]: b async for b in db.bookings.find({"id": {"$in": ids}}, {"_id": 0, "id": 1, "code": 1, "customer_name": 1})}
        for r in rows:
            b = bmap.get(r.get("booking_id"), {})
            r.setdefault("booking_code", b.get("code"))
            r.setdefault("customer_name", b.get("customer_name"))
    wb = Workbook()
    ws = wb.active
    ws.title = spec["label"][:31]
    head_font, head_fill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="0F6E56")
    cols = [c for c in spec["columns"] if any(r.get(c[0]) not in (None, "", [], 0) for r in rows)] or spec["columns"]
    for ci, (_, header) in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=header)
        c.font, c.fill, c.alignment = head_font, head_fill, Alignment(vertical="center")
    for ri, r in enumerate(rows, 2):
        for ci, (k, _) in enumerate(cols, 1):
            c = ws.cell(row=ri, column=ci, value=_cell(k, r.get(k)))
            if k in DATE_KEYS and isinstance(c.value, datetime):
                c.number_format = "DD/MM/YYYY HH:MM" if "datetime" in k or k in ("paid_at", "created_at") else "DD/MM/YYYY"
            elif k in MONEY_KEYS and isinstance(c.value, float):
                c.number_format = '"Rp" #,##0'
    for ci, (k, header) in enumerate(cols, 1):
        width = max([len(header)] + [min(len(str(r.get(k) or "")), 48) for r in rows[:500]])
        ws.column_dimensions[get_column_letter(ci)].width = max(10, min(width + 2, 50))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    buf = BytesIO()
    wb.save(buf)
    fname = f"{key}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return buf.getvalue(), fname, len(rows)


def table_catalog() -> list:
    out = []
    for k, v in TABLES.items():
        f = FILTERS.get(k)
        out.append({"key": k, "label": v["label"], "collection": v["collection"],
                    "filters": {"date_label": f["date_label"], "status": [{"key": sk, "label": sl, "options": opts} for sk, sl, opts in f["status"]]} if f else None})
    return out
